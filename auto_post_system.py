#!/usr/bin/env python3
"""
VET-ASSISTANT-CLI: 完全自動投稿システム
週間記事生成 → Excel出力 → 編集確認 → 毎朝7時自動投稿
"""

import os
import sys
import json
import argparse
import re
import schedule
import time
import threading
from datetime import datetime, timedelta
from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import tweepy
import google.generativeai as genai

class AutoPostSystem:
    def __init__(self):
        self.config_file = Path("config.json")
        self.persona_file = Path("persona_data.json")
        self.excel_file = Path("投稿スケジュール.xlsx")
        self.api_key = "AIzaSyAA0eEtEXToBEtZSrdllKJYZdkHQDrfgik"
        self.persona_data = {}
        self.twitter_api = None
        self.load_config()
        self.setup_gemini()
        
    def load_config(self):
        """設定ファイルを読み込み"""
        if self.config_file.exists():
            with open(self.config_file, 'r', encoding='utf-8') as f:
                self.config = json.load(f)
        else:
            self.config = {
                "x_archive_path": "",
                "learned": False,
                "current_week_type": "specialty",  # specialty, cat-breed, interactive
                "twitter_api_key": "",
                "twitter_api_secret": "",
                "twitter_access_token": "",
                "twitter_access_token_secret": "",
                "twitter_bearer_token": "",
                "scheduler_running": False
            }
            self.save_config()
    
    def save_config(self):
        """設定ファイルを保存"""
        with open(self.config_file, 'w', encoding='utf-8') as f:
            json.dump(self.config, f, ensure_ascii=False, indent=2)
    
    def setup_gemini(self):
        """Gemini APIの設定"""
        try:
            genai.configure(api_key=self.api_key)
            self.model = genai.GenerativeModel('gemini-pro')
        except:
            print("⚠️ Gemini APIが利用できません。簡易版で動作します。")
            self.model = None
    
    def setup_twitter_api(self):
        """Twitter APIの設定"""
        try:
            if not all([
                self.config.get("twitter_api_key"),
                self.config.get("twitter_api_secret"),
                self.config.get("twitter_access_token"),
                self.config.get("twitter_access_token_secret"),
                self.config.get("twitter_bearer_token")
            ]):
                print("❌ Twitter API認証情報が設定されていません")
                return False
            
            # Twitter API v2 クライアント
            self.twitter_api = tweepy.Client(
                bearer_token=self.config["twitter_bearer_token"],
                consumer_key=self.config["twitter_api_key"],
                consumer_secret=self.config["twitter_api_secret"],
                access_token=self.config["twitter_access_token"],
                access_token_secret=self.config["twitter_access_token_secret"],
                wait_on_rate_limit=True
            )
            
            # 認証テスト
            self.twitter_api.get_me()
            print("✅ Twitter API認証成功")
            return True
            
        except Exception as e:
            print(f"❌ Twitter API認証失敗: {e}")
            return False
    
    def load_persona(self):
        """ペルソナデータを読み込み"""
        if self.persona_file.exists():
            with open(self.persona_file, 'r', encoding='utf-8') as f:
                self.persona_data = json.load(f)
                return True
        print("❌ ペルソナデータが見つかりません。先にlearnコマンドを実行してください")
        return False
    
    def count_characters(self, text):
        """文字数カウント（改行除外）"""
        return len(text.replace('\n', ''))
    
    def get_week_schedule(self):
        """週間投稿スケジュールを決定"""
        week_type = self.config.get("current_week_type", "specialty")
        
        if week_type == "cat-breed":
            return {
                "mon": ("cat-breed", "概要紹介"),
                "tue": ("cat-breed", "歴史・起源"),
                "wed": ("cat-breed", "性格"),
                "thu": ("cat-breed", "身体的特徴"),
                "fri": ("cat-breed", "健康上の注意点"),
                "sat": ("cat-breed", "ケアのポイント"),
                "sun": ("interactive", "参加型投稿")
            }
        elif week_type == "specialty":
            return {
                "mon": ("specialty", "概要"),
                "tue": ("specialty", "原因・初期症状"),
                "wed": ("specialty", "進行症状・合併症"),
                "thu": ("specialty", "診断方法"),
                "fri": ("specialty", "治療・管理方法"),
                "sat": ("specialty", "家庭でのケア"),
                "sun": ("specialty", "予防・総括")
            }
        else:  # interactive
            return {
                "mon": ("interactive", "クイズ"),
                "tue": ("interactive", "クイズ解説"),
                "wed": ("interactive", "アンケート"),
                "thu": ("interactive", "アンケート結果"),
                "fri": ("interactive", "事例紹介"),
                "sat": ("specialty", "豆知識"),
                "sun": ("interactive", "体験談募集")
            }
    
    def generate_week_posts(self, start_date=None, base_topic=None):
        """1週間分の投稿を生成"""
        if not self.load_persona():
            return False
        
        if start_date is None:
            start_date = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
            # 次の月曜日から開始
            days_ahead = 0 - start_date.weekday()
            if days_ahead <= 0:
                days_ahead += 7
            start_date += timedelta(days=days_ahead)
        
        week_schedule = self.get_week_schedule()
        posts_data = []
        
        day_names = ["mon", "tue", "wed", "thu", "fri", "sat", "sun"]
        japanese_days = ["月曜日", "火曜日", "水曜日", "木曜日", "金曜日", "土曜日", "日曜日"]
        
        for i, day in enumerate(day_names):
            post_date = start_date + timedelta(days=i)
            post_type, theme = week_schedule[day]
            
            # トピックを決定
            if base_topic:
                if post_type == "cat-breed":
                    topic = f"{base_topic}（{theme}）"
                elif post_type == "specialty":
                    topic = f"{base_topic}（{theme}）"
                else:
                    topic = base_topic
            else:
                # デフォルトトピック
                if post_type == "cat-breed":
                    topic = f"アメリカンショートヘア（{theme}）"
                elif post_type == "specialty":
                    topic = f"猫の腎臓病（{theme}）"
                else:
                    topic = "猫の健康管理"
            
            # 投稿内容を生成
            content = self.generate_single_post(post_type, day, topic, japanese_days[i])
            char_count = self.count_characters(content)
            
            posts_data.append({
                "日付": post_date.strftime("%Y-%m-%d"),
                "曜日": japanese_days[i],
                "投稿タイプ": post_type,
                "テーマ": theme,
                "トピック": topic,
                "投稿内容": content,
                "文字数": char_count,
                "ステータス": "編集待ち",
                "投稿時間": "07:00"
            })
        
        return posts_data
    
    def generate_single_post(self, post_type, day, topic, day_jp):
        """単一投稿を生成"""
        if self.model:
            return self.generate_with_gemini(post_type, day, topic, day_jp)
        else:
            return self.generate_template_post(post_type, day, topic)
    
    def generate_with_gemini(self, post_type, day, topic, day_jp):
        """Geminiを使用して投稿生成"""
        prompt = f"""
あなたは19年目の救急獣医師@souhei1219です。以下の指示に従って投稿を生成してください。

## 絶対的なルール
1. 投稿は140文字以内（改行除く、絵文字・記号・ハッシュタグ含む）
2. 必ず「#猫のあれこれ」をハッシュタグとして含める
3. 「獣医師が教える！【トピック名】【絵文字】」の形式でタイトルを作成
4. 専門用語は高校生でも理解できる言葉に言い換える
5. 箇条書き（✅、💡、🐾など）を効果的に使用
6. 改行を使って読みやすくする

## 投稿条件
- 投稿タイプ: {post_type}
- 曜日: {day_jp}
- トピック: {topic}

## ペルソナデータ
{json.dumps(self.persona_data, ensure_ascii=False, indent=2)}

投稿文のみを出力してください。
"""
        
        try:
            response = self.model.generate_content(prompt)
            content = response.text.strip()
            
            # 文字数チェック
            if self.count_characters(content) > 140:
                # 短縮版を生成
                short_prompt = prompt + "\n\n※140文字を超過しました。より簡潔に生成してください。"
                response = self.model.generate_content(short_prompt)
                content = response.text.strip()
            
            return content
        except Exception as e:
            print(f"⚠️ Gemini生成エラー: {e}")
            return self.generate_template_post(post_type, day, topic)
    
    def generate_template_post(self, post_type, day, topic):
        """テンプレートベースの投稿生成"""
        emojis = self.persona_data.get("よく使う絵文字", ["🐱", "✅", "💡"])
        main_emoji = emojis[0] if emojis else "🐱"
        
        if post_type == "specialty":
            content = f"""獣医師が教える！【{topic}】{main_emoji}

{topic}について大切なポイントをお話しします。

✅ 早期発見が重要
💡 適切な対処法
{main_emoji} 愛猫の健康を守りましょう

#猫のあれこれ"""
        
        elif post_type == "cat-breed":
            content = f"""獣医師が教える！【{topic}】{main_emoji}

{topic}の特徴について：

✅ 性格と特徴
💡 健康管理のポイント
{main_emoji} 飼育のコツ

#猫のあれこれ"""
        
        elif post_type == "interactive":
            content = f"""獣医師が教える！【{topic}】{main_emoji}

みなさんの愛猫はどうですか？

✅ こんな経験ありませんか？
💡 皆さんの体験談をお聞かせください
{main_emoji} コメントでお待ちしています

#猫のあれこれ"""
        
        return content
    
    def save_to_excel(self, posts_data):
        """Excelファイルに保存"""
        try:
            df = pd.DataFrame(posts_data)
            
            # Excelファイルに書き込み
            with pd.ExcelWriter(self.excel_file, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='投稿スケジュール', index=False)
                
                # スタイル設定
                workbook = writer.book
                worksheet = writer.sheets['投稿スケジュール']
                
                # ヘッダースタイル
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)
                
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center")
                
                # 列幅調整
                column_widths = {
                    'A': 12,  # 日付
                    'B': 8,   # 曜日
                    'C': 12,  # 投稿タイプ
                    'D': 15,  # テーマ
                    'E': 20,  # トピック
                    'F': 50,  # 投稿内容
                    'G': 8,   # 文字数
                    'H': 12,  # ステータス
                    'I': 10   # 投稿時間
                }
                
                for col, width in column_widths.items():
                    worksheet.column_dimensions[col].width = width
                
                # 投稿内容列の折り返し設定
                for row in worksheet.iter_rows(min_row=2, max_row=len(posts_data)+1):
                    row[5].alignment = Alignment(wrap_text=True, vertical="top")
            
            print(f"✅ Excelファイルに保存しました: {self.excel_file}")
            return True
            
        except Exception as e:
            print(f"❌ Excel保存エラー: {e}")
            return False
    
    def read_excel_schedule(self):
        """Excelファイルから投稿スケジュールを読み込み"""
        try:
            if not self.excel_file.exists():
                print("❌ 投稿スケジュールファイルが見つかりません")
                return []
            
            df = pd.read_excel(self.excel_file, sheet_name='投稿スケジュール')
            return df.to_dict('records')
            
        except Exception as e:
            print(f"❌ Excel読み込みエラー: {e}")
            return []
    
    def post_to_twitter(self, content):
        """Twitterに投稿"""
        try:
            if not self.twitter_api:
                if not self.setup_twitter_api():
                    return False
            
            response = self.twitter_api.create_tweet(text=content)
            print(f"✅ 投稿成功: {response.data['id']}")
            return True
            
        except Exception as e:
            print(f"❌ 投稿エラー: {e}")
            return False
    
    def check_and_post(self):
        """投稿チェックと実行"""
        current_time = datetime.now()
        today = current_time.strftime("%Y-%m-%d")
        current_hour = current_time.hour
        
        # 7時でない場合はスキップ
        if current_hour != 7:
            return
        
        print(f"🔍 投稿チェック中... ({today} {current_time.strftime('%H:%M')})")
        
        schedule_data = self.read_excel_schedule()
        if not schedule_data:
            return
        
        for row in schedule_data:
            # 今日の投稿をチェック
            if (row.get("日付") == today and 
                "#猫のあれこれ" in str(row.get("投稿内容", "")) and
                row.get("ステータス") != "投稿済み"):
                
                content = row.get("投稿内容", "")
                print(f"📝 投稿実行: {content[:30]}...")
                
                if self.post_to_twitter(content):
                    # ステータスを更新
                    self.update_post_status(row, "投稿済み")
                    print("✅ 投稿完了")
                else:
                    print("❌ 投稿失敗")
    
    def update_post_status(self, row_data, status):
        """投稿ステータスを更新"""
        try:
            schedule_data = self.read_excel_schedule()
            for i, row in enumerate(schedule_data):
                if (row.get("日付") == row_data.get("日付") and 
                    row.get("投稿内容") == row_data.get("投稿内容")):
                    schedule_data[i]["ステータス"] = status
                    break
            
            # Excelファイルを更新
            df = pd.DataFrame(schedule_data)
            df.to_excel(self.excel_file, sheet_name='投稿スケジュール', index=False)
            
        except Exception as e:
            print(f"❌ ステータス更新エラー: {e}")
    
    def start_scheduler(self):
        """スケジューラーを開始"""
        print("🚀 自動投稿スケジューラーを開始します...")
        print("⏰ 毎朝7時に投稿をチェックします")
        
        # 毎時チェック（7時のみ実行）
        schedule.every().hour.do(self.check_and_post)
        
        self.config["scheduler_running"] = True
        self.save_config()
        
        def run_scheduler():
            while self.config.get("scheduler_running", False):
                schedule.run_pending()
                time.sleep(60)  # 1分ごとにチェック
        
        # バックグラウンドで実行
        scheduler_thread = threading.Thread(target=run_scheduler, daemon=True)
        scheduler_thread.start()
        
        print("✅ スケジューラー開始完了")
        print("💡 停止するには Ctrl+C を押してください")
        
        try:
            while True:
                time.sleep(1)
        except KeyboardInterrupt:
            print("\n🛑 スケジューラーを停止します...")
            self.config["scheduler_running"] = False
            self.save_config()

def main():
    parser = argparse.ArgumentParser(description="VET-ASSISTANT-CLI: 完全自動投稿システム")
    subparsers = parser.add_subparsers(dest='command', help='利用可能なコマンド')
    
    # generate-weekコマンド
    week_parser = subparsers.add_parser('generate-week', help='1週間分の投稿を生成')
    week_parser.add_argument('--topic', help='基本トピック（例: 猫の腎臓病）')
    week_parser.add_argument('--start-date', help='開始日（YYYY-MM-DD）')
    
    # start-schedulerコマンド
    scheduler_parser = subparsers.add_parser('start-scheduler', help='自動投稿スケジューラーを開始')
    
    # setup-twitterコマンド
    twitter_parser = subparsers.add_parser('setup-twitter', help='Twitter API設定')
    
    # test-postコマンド
    test_parser = subparsers.add_parser('test-post', help='投稿テスト')
    test_parser.add_argument('content', help='投稿内容')
    
    args = parser.parse_args()
    
    if not args.command:
        parser.print_help()
        return
    
    app = AutoPostSystem()
    
    if args.command == 'generate-week':
        start_date = None
        if args.start_date:
            start_date = datetime.strptime(args.start_date, "%Y-%m-%d")
        
        print("📝 1週間分の投稿を生成中...")
        posts_data = app.generate_week_posts(start_date, args.topic)
        
        if posts_data:
            app.save_to_excel(posts_data)
            print("🎉 週間投稿生成完了！")
            print(f"📊 {app.excel_file} を確認して編集してください")
        else:
            print("❌ 投稿生成に失敗しました")
    
    elif args.command == 'start-scheduler':
        app.start_scheduler()
    
    elif args.command == 'setup-twitter':
        print("🔐 Twitter API設定")
        app.config["twitter_api_key"] = input("API Key: ")
        app.config["twitter_api_secret"] = input("API Secret: ")
        app.config["twitter_access_token"] = input("Access Token: ")
        app.config["twitter_access_token_secret"] = input("Access Token Secret: ")
        app.config["twitter_bearer_token"] = input("Bearer Token: ")
        app.save_config()
        
        if app.setup_twitter_api():
            print("✅ Twitter API設定完了")
        else:
            print("❌ 設定に失敗しました")
    
    elif args.command == 'test-post':
        print(f"🧪 投稿テスト: {args.content}")
        if app.post_to_twitter(args.content):
            print("✅ テスト投稿成功")
        else:
            print("❌ テスト投稿失敗")

if __name__ == "__main__":
    main()